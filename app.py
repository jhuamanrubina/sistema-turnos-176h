import streamlit as st
import pandas as pd
import os
import datetime
import calendar
import random

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from reportlab.platypus import SimpleDocTemplate, Table

# ================= CONFIGURACIÓN =================
COORDINADORES_AUTORIZADOS = {
    "Samay02": "pass123",
    "Yape": "yape2024",
    "Capacity": "capa123",
    "Samay01": "pass123",
    "Admin": "admin789"
}

DB_FILE = "especialistas_vFinal.csv"
TURNOS_ESPECIALES_FILE = "turnos_especiales.csv"

TURNOS_OPCIONES = ["6am-2pm", "9am-6pm", "6pm-2am", "10pm-6am"]
POOLS_DISPONIBLES = ["Samay02", "Yape", "proyectos", "Legacy", "Samay01", "SYF", "Capacity"]

# ================= DATA =================
def cargar_datos():
    if os.path.exists(DB_FILE):
        df = pd.read_csv(DB_FILE)
        if 'Turno_Fijo' not in df.columns:
            df['Turno_Fijo'] = "Aleatorio"
        if 'Prestado_A' not in df.columns:
            df['Prestado_A'] = ""
        return df
    return pd.DataFrame(columns=['Nombre','Pool','Coordinador','Turno_Fijo','Prestado_A'])

def guardar_datos(df):
    df.to_csv(DB_FILE, index=False)

def cargar_turnos_especiales():
    if os.path.exists(TURNOS_ESPECIALES_FILE):
        return pd.read_csv(TURNOS_ESPECIALES_FILE)
    return pd.DataFrame(columns=['Coordinador','Especialista','Día','Turno'])

def guardar_turnos_especiales(df):
    df.to_csv(TURNOS_ESPECIALES_FILE, index=False)

# ================= GENERADOR =================
def generar_rol_perfecto(mes, anio, df_base, coord):
    df_manual = cargar_turnos_especiales()
    num_dias = calendar.monthrange(anio, mes)[1]

    df_filt = df_base[df_base['Coordinador'] == coord]
    especialistas = df_filt['Nombre'].tolist()
    if not especialistas:
        return pd.DataFrame(), {}

    patron = ["6am-2pm","9am-6pm","9am-6pm","6pm-2am","10pm-6am"]
    mapa_turnos = {
        n: (
            df_filt[df_filt['Nombre']==n]['Turno_Fijo'].values[0]
            if df_filt[df_filt['Nombre']==n]['Turno_Fijo'].values[0] in TURNOS_OPCIONES
            else patron[i % len(patron)]
        )
        for i,n in enumerate(especialistas)
    }

    asignaciones = []
    horas = {n:0 for n in especialistas}

    for dia in range(1, num_dias+1):
        random.shuffle(especialistas)
        for n in especialistas:
            dias_trab = sorted([a['Día'] for a in asignaciones if a['Especialista']==n])
            if dia in dias_trab:
                continue
            if horas[n] >= 176:
                continue
            if len([d for d in dias_trab if dia-d <= 6]) >= 6:
                continue

            manual = df_manual[
                (df_manual['Coordinador']==coord) &
                (df_manual['Especialista']==n) &
                (df_manual['Día']==dia)
            ]

            turno = manual['Turno'].values[0] if not manual.empty else mapa_turnos[n]

            asignaciones.append({
                "Día":dia,
                "Especialista":n,
                "Turno":turno,
                "Pool":df_filt[df_filt['Nombre']==n]['Pool'].values[0]
            })
            horas[n] += 8

    # Ajuste final para llegar a 176h respetando descanso
    for n in especialistas:
        while horas[n] < 176:
            for dia in range(1, num_dias+1):
                dias_trab = sorted([a['Día'] for a in asignaciones if a['Especialista']==n])
                if dia in dias_trab:
                    continue
                if len([d for d in dias_trab if dia-d <= 6]) >= 6:
                    continue

                asignaciones.append({
                    "Día":dia,
                    "Especialista":n,
                    "Turno":mapa_turnos[n],
                    "Pool":df_filt[df_filt['Nombre']==n]['Pool'].values[0]
                })
                horas[n] += 8
                break

    return pd.DataFrame(asignaciones), horas

# ================= EXPORTACIONES =================
def exportar_excel(df):
    wb = Workbook()
    ws = wb.active

    colores = {
        "6am-2pm":"D1E9F6","9am-6pm":"FFF9BF",
        "6pm-2am":"F1D3FF","10pm-6am":"D1FFD7",
        "DESCANSO":"FFD1D1"
    }

    ws.append(["Especialista"]+list(df.columns))
    for esp,row in df.iterrows():
        ws.append([esp]+list(row.values))

    for r in ws.iter_rows(min_row=2,min_col=2):
        for c in r:
            if c.value in colores:
                c.fill = PatternFill(start_color=colores[c.value], end_color=colores[c.value], fill_type="solid")

    wb.save("rol_mensual.xlsx")

def exportar_pdf(df):
    pdf = SimpleDocTemplate("rol_mensual.pdf")
    data = [["Especialista"]+list(df.columns)]
    for esp,row in df.iterrows():
        data.append([esp]+list(row.values))
    pdf.build([Table(data)])

# ================= UI =================
st.set_page_config("Control 176h", layout="wide")

u = st.sidebar.selectbox("Usuario", COORDINADORES_AUTORIZADOS.keys())
p = st.sidebar.text_input("Contraseña", type="password")

if p != COORDINADORES_AUTORIZADOS.get(u):
    st.info("Credenciales requeridas")
    st.stop()

if u == "Capacity":
    st.warning("Capacity no genera roles")
    st.stop()

df_base = cargar_datos()

t1,t2,t3 = st.tabs(["🗓️ Rol","👥 Personal","📊 Auditoría"])

# -------- PERSONAL --------
with t2:
    st.subheader("Capacity exclusivo")
    cap = df_base[(df_base['Pool']=="Capacity") & (df_base['Prestado_A']=="")]
    if not cap.empty:
        sel = st.selectbox("Seleccionar", cap['Nombre'])
        if st.button("Asignar"):
            df_base.loc[df_base['Nombre']==sel,['Prestado_A','Coordinador']] = [u,u]
            guardar_datos(df_base)
            st.rerun()

    st.subheader("Mi equipo")
    st.dataframe(df_base[df_base['Coordinador']==u][['Nombre','Pool','Turno_Fijo']])

# -------- ROL --------
with t1:
    st.subheader("Cambio manual por persona y día")
    esp = st.selectbox("Especialista", df_base[df_base['Coordinador']==u]['Nombre'])
    d = st.number_input("Día",1,31)
    tr = st.selectbox("Turno", TURNOS_OPCIONES)

    if st.button("Aplicar cambio"):
        df_m = cargar_turnos_especiales()
        df_m = df_m[~((df_m['Coordinador']==u)&(df_m['Especialista']==esp)&(df_m['Día']==d))]
        df_m = pd.concat([df_m,pd.DataFrame([[u,esp,d,tr]],columns=df_m.columns)])
        guardar_turnos_especiales(df_m)
        st.success("Cambio aplicado")

    mes = st.selectbox("Mes", range(1,13), datetime.datetime.now().month-1)
    if st.button("Generar rol"):
        r,h = generar_rol_perfecto(mes,2026,df_base,u)
        st.session_state['r']=r
        st.session_state['h']=h

    if 'r' in st.session_state:
        mat = st.session_state['r'].pivot(index='Especialista',columns='Día',values='Turno').fillna("DESCANSO")

        def color_turnos(v):
            return f"background-color:{ {'6am-2pm':'#D1E9F6','9am-6pm':'#FFF9BF','6pm-2am':'#F1D3FF','10pm-6am':'#D1FFD7','DESCANSO':'#FFD1D1'}.get(v,'white')}"

        st.dataframe(mat.style.applymap(color_turnos), use_container_width=True)

        exportar_excel(mat)
        exportar_pdf(mat)

        with open("rol_mensual.xlsx","rb") as f:
            st.download_button("📤 Excel",f,"rol_mensual.xlsx")

        with open("rol_mensual.pdf","rb") as f:
            st.download_button("📄 PDF",f,"rol_mensual.pdf")

# -------- AUDITORÍA --------
with t3:
    if 'h' in st.session_state:
        df_h = pd.DataFrame([{"Especialista":k,"Horas":v} for k,v in st.session_state['h'].items()])

        def color_h(v):
            if v==176: return "background-color:#2ecc71;color:white"
            if v<176: return "background-color:#f1c40f"
            return "background-color:#e74c3c;color:white"

        st.dataframe(df_h.style.applymap(color_h,subset=['Horas']))
